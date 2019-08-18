#!/usr/bin/env python3
# coding=utf-8
"""
Copyright 2019 SHIGENO Yoshitaka <shigeno@coop.nagoya-u.ac.jp>

Copying and distribution of this file, with or without modification,
are permitted in any medium without royalty provided the copyright
notice and this notice are preserved.  This file is offered as-is,
without any warranty.
"""

import argparse
import csv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule

if(__name__ == "__main__"):
    parser = argparse.ArgumentParser()
    parser.add_argument('people',
                        type=argparse.FileType('r'),
                        help='People CSV')
    parser.add_argument('works',
                        type=argparse.FileType('r'),
                        help='Works CSV.')
    parser.add_argument('assignment',
                        type=argparse.FileType('r'),
                        help='Assignment result CSV.')
    args = parser.parse_args()

    people = []
    people_csv = csv.reader(args.people)
    for row in people_csv:
        tmp = []
        for c in row:
            val = None
            try:
                val = int(c)
            except ValueError:
                val = str(c)
            tmp.append(val)
        people.append(tmp)
    people.sort()

    works = []
    works_csv = csv.reader(args.works)
    for row in works_csv:
        tmp = []
        for c in row:
            val = None
            try:
                val = int(c)
            except ValueError:
                val = str(c)
            tmp.append(val)
        works.append(tmp)
    works.sort()

    assignment = []
    assigment_csv = csv.reader(args.assignment)
    for row in assigment_csv:
        tmp = []
        for c in row:
            val = None
            try:
                val = int(c)
            except Exception:
                val = c
            tmp.append(val)
        assignment.append(tmp)
    assignment.sort()

    wb = Workbook()
    ws = wb.active
    ws.title = "People"
    ws.append(["pid", "white_list", "time"])
    for row in people:
        ws.append(row)
    table_people = Table(displayName="people", ref="A1:C"+str(len(people)+1))
    ws.add_table(table_people)

    ws = wb.create_sheet("Works")
    ws.append(["wid", "type", "label", "time"])
    ws.append([-1, "other", ""])
    ws.append([0, "absent", ""])
    for row in works:
        ws.append(row)
    table_works = Table(displayName="works", ref="A1:D"+str(len(works)+3))
    ws.add_table(table_works)

    ws = wb.create_sheet("Assignment")
    ws.append(["pid", "9:00", "9:30", "10:00", "10:30", "11:00", "11:30",
               "12:00", "12:30", "13:00", "13:30", "14:00", "14:30", "15:00",
               "15:30", "16:00", "16:30", "17:00"])
    for row in assignment:
        ws.append(row)
    table_assignment = Table(displayName="assignment",
                             ref="A1:R"+str(len(assignment)+1))
    ws.add_table(table_assignment)

    ws = wb.create_sheet("Shift")
    ws.append(["pid", "9:00", "9:30", "10:00", "10:30", "11:00", "11:30",
               "12:00", "12:30", "13:00", "13:30", "14:00", "14:30", "15:00",
               "15:30", "16:00", "16:30", "17:00"])
    for row in people:
        ws.append(
            [row[0], '=VLOOKUP(-1*VLOOKUP($A2,assignment,COLUMN()),works,3)'])

    ws.conditional_formatting.add("A2:R"+str(len(people)+1), FormulaRule(
        formula=[
            '=EXACT(VLOOKUP(-1*VLOOKUP($A2,' +
            "'Assignment'!$A$2:$R$" + str(len(assignment)+1) +
            ',COLUMN()),' +
            "'Works'!$A$2:$D$"+str(len(works)+3) +
            ',2),"absent")'],
        stopIfTrue=True,
        fill=PatternFill(patternType='solid',
                         fgColor='FF000000', bgColor='FF000000')
    ))
    ws.conditional_formatting.add("A2:R"+str(len(people)+1), FormulaRule(
        formula=[
            '=EXACT(VLOOKUP(-1*VLOOKUP($A2,' +
            "'Assignment'!$A$2:$R$" + str(len(assignment)+1) +
            ',COLUMN()),' +
            "'Works'!$A$2:$D$"+str(len(works)+3) +
            ',2),"parasol")'],
        stopIfTrue=True,
        fill=PatternFill(patternType='solid',
                         fgColor='FFFFDDDD', bgColor='FFFFDDDD')
    ))
    ws.conditional_formatting.add("A2:R"+str(len(people)+1), FormulaRule(
        formula=[
            '=EXACT(VLOOKUP(-1*VLOOKUP($A2,' +
            "'Assignment'!$A$2:$R$" + str(len(assignment)+1) +
            ',COLUMN()),' +
            "'Works'!$A$2:$D$"+str(len(works)+3) +
            ',2),"tour")'],
        stopIfTrue=True,
        fill=PatternFill(patternType='solid',
                         fgColor='FFDDFFDD', bgColor='FFDDFFDD')
    ))
    ws.conditional_formatting.add("A2:R"+str(len(people)+1), FormulaRule(
        formula=[
            '=EXACT(VLOOKUP(-1*VLOOKUP($A2,' +
            "'Assignment'!$A$2:$R$" + str(len(assignment)+1) +
            ',COLUMN()),' +
            "'Works'!$A$2:$D$"+str(len(works)+3) +
            ',2),"sc")'],
        stopIfTrue=True,
        fill=PatternFill(patternType='solid',
                         fgColor='FFDDDDFF', bgColor='FFDDDDFF')
    ))
    ws.conditional_formatting.add("A2:R"+str(len(people)+1), FormulaRule(
        formula=[
            '=EXACT(VLOOKUP(-1*VLOOKUP($A2,' +
            "'Assignment'!$A$2:$R$" + str(len(assignment)+1) +
            ',COLUMN()),' +
            "'Works'!$A$2:$D$"+str(len(works)+3) +
            ',2),"named")'],
        stopIfTrue=True,
        fill=PatternFill(patternType='solid',
                         fgColor='FFDDDDDD', bgColor='FFDDDDDD')
    ))

    wb.save("sample.xlsx")
